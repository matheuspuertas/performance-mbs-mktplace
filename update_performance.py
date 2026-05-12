#!/usr/bin/env python3
"""
Atualizador de Performance MBS MKTPLACE
Le relatorios do ML e Shopee e atualiza a planilha de performance.
"""

import os
import re
import sys
import json
import base64
import shutil
from copy import copy
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd

# -------------------------------------------------------------------------------
# CONFIGURACOES
# -------------------------------------------------------------------------------

PERFORMANCE_FILE = r"P:\Meu Drive\Empresas\MBS Pro Grooming\Performance MBS MKTPLACE.xlsx"
MATERIAIS_DIR    = r"P:\Meu Drive\Empresas\MBS Pro Grooming\materiais"

MESES_PT = {
    1: "jan", 2: "fev", 3: "mar",  4: "abr",
    5: "mai", 6: "jun", 7: "jul",  8: "ago",
    9: "set",10: "out",11: "nov", 12: "dez",
}

# Indices de coluna de cada aba (1-based)
ML_COLS = {
    "mes": 1,
    # Atividade dos compradores
    "visitas":                  2,
    "compradores_unicos":       3,
    "media_vendas_comprador":   4,
    "novos_compradores":        5,
    "compradores_existentes":   6,
    "taxa_recompra":            7,
    # Desempenho de vendas
    "qtd_vendas":               8,
    "unidades_vendidas":        9,
    "vendas_brutas":           10,
    "conversao":               11,
    "valor_medio_venda":       12,
    "preco_medio_unidade":     13,
    # Publicidade
    "pub_impressoes":          14,
    "pub_cliques":             15,
    "pub_vendas_product_ads":  16,
    "pub_vendas_sem_products": 17,
    "pub_investimento":        18,
    "pub_receita":             19,
    "pub_acos":                20,   # formula
    "pub_roas":                21,   # formula
    "pub_tacos":               22,   # formula
    "pub_fat_total":           23,   # formula
    # Afiliados
    "af_receita":              24,
    "af_unidades":             25,
    "af_qtd_vendas":           26,
    "af_custo":                27,
    "af_ticket_medio":         28,   # formula
    "af_acos":                 29,   # formula
    "af_fat_total":            30,   # formula
    "af_rec_organica":         31,   # formula
}
ML_FORMULA_COLS = {20, 21, 22, 23, 28, 29, 30, 31}
ML_LINHA_DADOS  = 3

SHOPEE_COLS = {
    "mes": 1,
    # Dados gerais
    "visitantes":              2,
    "num_pedidos":             3,
    "pedidos_feitos_valor":    4,
    "pedidos_pagos_qtd":       5,
    "pedidos_pagos_valor":     6,
    "ticket_medio":            7,
    "taxa_conv_feitos":        8,
    "taxa_conv_pagos":         9,
    "taxa_conv_feitos_pagos": 10,
    # Publicidade
    "pub_impressoes":         11,
    "pub_cliques":            12,
    "pub_ctr":                13,
    "pub_conversoes":         14,
    "pub_conv_diretas":       15,
    "pub_taxa_conv":          16,
    "pub_taxa_conv_direta":   17,
    "pub_itens_vendidos":     18,
    "pub_itens_diretos":      19,
    "pub_gmv":                20,
    "pub_receita_ads":        21,
    "pub_despesas":           22,
    "pub_roas":               23,
    "pub_roas_direto":        24,
    "pub_acos":               25,
    "pub_acos_direto":        26,
    "pub_tacos":              27,   # formula
    "pub_fat_total":          28,   # formula
    # Afiliados
    "af_vendas":              29,
    "af_itens_brutos":        30,
    "af_pedidos":             31,
    "af_cliques":             32,
    "af_comissao":            33,
    "af_roi":                 34,
    "af_compradores_totais":  35,
    "af_novos_compradores":   36,
    "af_fat_total":           37,   # formula
    "af_rec_organica":        38,   # formula
}
SHOPEE_FORMULA_COLS = {27, 28, 37, 38}
SHOPEE_LINHA_DADOS  = 3

SITE_COLS = {
    "mes":            1,
    "usuarios":       2,
    "num_pedidos":    3,
    "valor_total":    4,
    "ticket_medio":   5,
    "valor_investido":6,
    "google":         7,
    "meta_ads":       8,
    "taxa_tacos":     9,
    "taxa_conversao": 10,
}
SITE_FORMULA_COLS = set()
SITE_LINHA_DADOS  = 4


# -------------------------------------------------------------------------------
# API KEY
# -------------------------------------------------------------------------------

def carregar_api_key():
    """Carrega a chave da Anthropic do config.env ou variavel de ambiente."""
    config = Path(__file__).parent / "config.env"
    if config.exists():
        for line in config.read_text().splitlines():
            if line.startswith("ANTHROPIC_API_KEY="):
                return line.split("=", 1)[1].strip()
    return os.environ.get("ANTHROPIC_API_KEY")


# -------------------------------------------------------------------------------
# VISAO COM IA (Claude)
# -------------------------------------------------------------------------------

PROMPT_VISAO = """Analise esta captura de tela de um painel de marketing e identifique o tipo:

- "ml_ads": Publicidade Mercado Livre (tem "Vendas por Product Ads" e grafico de barras azul)
- "ml_af": Afiliados Mercado Livre (tem abas "Rendimento geral", "Metricas de campanhas", "Metricas de afiliados")
- "shopee_af": Afiliados Shopee (tem "Principais Indicadores", abas "Shopee Live", "Shopee Video")

Retorne APENAS um JSON (sem texto adicional):

Se ml_ads:   {"tipo":"ml_ads","impressoes":N,"cliques":N,"vendas_product_ads":N,"vendas_sem_products":N,"investimento":N.NN,"receita":N.NN}
Se ml_af:    {"tipo":"ml_af","receita":N.NN,"unidades":N,"qtd_vendas":N,"custo":N.NN}
Se shopee_af:{"tipo":"shopee_af","vendas":N.NN,"itens_brutos":N,"pedidos":N,"cliques":N,"comissao":N.NN,"roi":N.N,"compradores_totais":N,"novos_compradores":N}

Regras importantes:
- Valores monetarios: apenas o numero sem R$ (ex: 22314.47)
- "70mil" ou "R$70mil" = 70000.0
- "14,8mil" = 14800
- "3,6mil" = 3600.0
- "11.332.096" (formato BR com pontos) = 11332096
- Nao incluir percentuais no JSON"""

CAMPOS_IA = {
    "ml_ads": {
        "impressoes":         "pub_impressoes",
        "cliques":            "pub_cliques",
        "vendas_product_ads": "pub_vendas_product_ads",
        "vendas_sem_products":"pub_vendas_sem_products",
        "investimento":       "pub_investimento",
        "receita":            "pub_receita",
    },
    "ml_af": {
        "receita":    "af_receita",
        "unidades":   "af_unidades",
        "qtd_vendas": "af_qtd_vendas",
        "custo":      "af_custo",
    },
    "shopee_af": {
        "vendas":             "af_vendas",
        "itens_brutos":       "af_itens_brutos",
        "pedidos":            "af_pedidos",
        "cliques":            "af_cliques",
        "comissao":           "af_comissao",
        "roi":                "af_roi",
        "compradores_totais": "af_compradores_totais",
        "novos_compradores":  "af_novos_compradores",
    },
}

LABELS_TIPO = {
    "ml_ads":    "ML -- Publicidade",
    "ml_af":     "ML -- Afiliados",
    "shopee_af": "Shopee -- Afiliados",
}

LABELS_CAMPO = {
    "impressoes": "Impressoes", "cliques": "Cliques",
    "vendas_product_ads": "Vendas Product Ads", "vendas_sem_products": "Vendas sem Products",
    "investimento": "Investimento (R$)", "receita": "Receita (R$)",
    "unidades": "Unidades vendidas", "qtd_vendas": "Qtd vendas", "custo": "Custo estimado (R$)",
    "vendas": "Vendas (R$)", "itens_brutos": "Itens brutos", "pedidos": "Pedidos",
    "comissao": "Comissao (R$)", "roi": "ROI",
    "compradores_totais": "Compradores totais", "novos_compradores": "Novos compradores",
}


def extrair_dados_imagem_claude(caminho, api_key):
    """Envia imagem para o Claude e retorna dicionario com tipo e valores."""
    import anthropic

    ext = caminho.suffix.lower()
    media_type = "image/jpeg" if ext in (".jpg", ".jpeg") else "image/png"

    with open(str(caminho), "rb") as f:
        img_b64 = base64.b64encode(f.read()).decode("utf-8")

    client = anthropic.Anthropic(api_key=api_key)
    resp = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=512,
        messages=[{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": img_b64}},
                {"type": "text",  "text": PROMPT_VISAO},
            ],
        }],
    )

    texto = resp.content[0].text.strip()
    m = re.search(r"\{.*\}", texto, re.DOTALL)
    if m:
        return json.loads(m.group())
    return None


def processar_imagens_ia(imagens, api_key):
    """Processa todas as imagens com IA e retorna dados para ML e Shopee."""
    dados_ml     = {}
    dados_shopee = {}

    for img in sorted(imagens):
        print(f"\n  Analisando: {img.name} ...")
        try:
            resultado = extrair_dados_imagem_claude(img, api_key)
        except Exception as e:
            print(f"  [!] Erro na IA: {e}")
            resultado = None

        if not resultado or "tipo" not in resultado:
            print("  [!] Nao foi possivel ler automaticamente. Entrada manual:")
            _fallback_manual(img, dados_ml, dados_shopee)
            continue

        tipo = resultado["tipo"]
        print(f"  [OK] Tipo identificado: {LABELS_TIPO.get(tipo, tipo)}")
        print("  Valores extraidos:")
        for chave_ia in CAMPOS_IA.get(tipo, {}):
            val = resultado.get(chave_ia)
            if val is not None:
                label = LABELS_CAMPO.get(chave_ia, chave_ia)
                s = f"{val:,.2f}" if isinstance(val, float) else f"{val:,}"
                print(f"    {label}: {s}")

        resp = input("  Confirmar? [S/n/e(ditar)]: ").strip().lower()

        if resp == "n":
            print("  Pulando.")
            continue

        if resp == "e":
            for chave_ia in CAMPOS_IA.get(tipo, {}):
                val_atual = resultado.get(chave_ia, "")
                novo = input(f"    {LABELS_CAMPO.get(chave_ia, chave_ia)} [{val_atual}]: ").strip()
                if novo:
                    resultado[chave_ia] = limpar_numero(novo)

        # Mapear para colunas da planilha
        destino = dados_ml if tipo in ("ml_ads", "ml_af") else dados_shopee
        for chave_ia, chave_col in CAMPOS_IA.get(tipo, {}).items():
            if chave_ia in resultado and resultado[chave_ia] is not None:
                destino[chave_col] = resultado[chave_ia]

    return dados_ml, dados_shopee


def _fallback_manual(img, dados_ml, dados_shopee):
    """Fallback de entrada manual quando a IA falha."""
    abrir_imagem(img)
    tipos = {
        "1": ("ML -- Publicidade", "ml_ads"),
        "2": ("ML -- Afiliados",   "ml_af"),
        "3": ("Shopee -- Afiliados","shopee_af"),
        "0": ("Ignorar", None),
    }
    for k, (desc, _) in tipos.items():
        print(f"  [{k}] {desc}")
    while True:
        op = input("  Tipo: ").strip()
        if op in tipos:
            break
    tipo = tipos[op][1]
    if not tipo:
        return

    if tipo == "ml_ads":
        dados_ml.update(coletar_ml_ads(img))
    elif tipo == "ml_af":
        dados_ml.update(coletar_ml_afiliados(img))
    elif tipo == "shopee_af":
        dados_shopee.update(coletar_shopee_afiliados(img))


# -------------------------------------------------------------------------------
# HELPERS
# -------------------------------------------------------------------------------

def limpar_numero(texto):
    """Converte string de numero (formato BR ou simples) para float."""
    if texto is None:
        return None
    s = str(texto).strip().replace("R$", "").replace(" ", "").replace("%", "")
    if not s:
        return None
    # "70mil" -> 70000 | "3,6mil" -> 3600
    m = re.match(r"^([\d.,]+)mil$", s, re.IGNORECASE)
    if m:
        num = m.group(1).replace(".", "").replace(",", ".")
        try:
            return float(num) * 1000
        except ValueError:
            pass
    # Formato BR: 1.234,56 -> 1234.56
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def mes_serial_excel(ano, mes):
    """Retorna o numero serial Excel para o dia 1 do mes."""
    d = date(ano, mes, 1)
    base = date(1899, 12, 30)
    return (d - base).days


def ajustar_formula(formula, delta=1):
    """Incrementa referencias de linha relativas numa formula Excel."""
    if not formula or not str(formula).startswith("="):
        return formula

    def _sub(m):
        col = m.group(1)
        row = int(m.group(2))
        return f"{col}{row + delta}"

    return re.sub(r"(?<!\$)([A-Z]+)(\d+)", _sub, str(formula))


def copiar_estilo(origem, destino):
    """Copia formatacao de celula (incluindo formato de numero)."""
    if origem.has_style:
        destino.font          = copy(origem.font)
        destino.border        = copy(origem.border)
        destino.fill          = copy(origem.fill)
        destino.number_format = origem.number_format
        destino.alignment     = copy(origem.alignment)


def input_numero(prompt, obrigatorio=True):
    """Le numero do usuario com validacao."""
    while True:
        raw = input(prompt).strip()
        if not raw and not obrigatorio:
            return None
        if not raw:
            print("  [!] Campo obrigatorio.")
            continue
        val = limpar_numero(raw)
        if val is not None:
            return val
        print("  [!] Valor invalido. Exemplos: 1234  |  1.234,56  |  22.314,47  |  70mil")


def abrir_imagem(caminho):
    """Abre imagem no visualizador padrao do Windows."""
    os.startfile(str(caminho))


# -------------------------------------------------------------------------------
# DETECCAO DE MES
# -------------------------------------------------------------------------------

def detectar_mes(pasta):
    """Detecta mes/ano pelos nomes dos arquivos na pasta."""
    encontrados = []
    for arq in Path(pasta).iterdir():
        nome = arq.name
        # ML: Relatorio_evolucao_negocio_2026_04_01-2026_04_30.xlsx
        m = re.search(r"(\d{4})_(\d{2})_\d{2}-\d{4}_\d{2}_\d{2}", nome)
        if m:
            encontrados.append((int(m.group(1)), int(m.group(2))))
            continue
        # Shopee sales_overview_20260401-20260430.xlsx
        m = re.search(r"(\d{4})(\d{2})\d{2}-\d{4}\d{2}\d{2}", nome)
        if m:
            encontrados.append((int(m.group(1)), int(m.group(2))))
            continue
        # Shopee CSV: Dados+Gerais-01_04_2026-30_04_2026.csv
        m = re.search(r"(\d{2})_(\d{2})_(\d{4})-\d{2}_\d{2}_\d{4}", nome)
        if m:
            encontrados.append((int(m.group(3)), int(m.group(2))))
            continue

    if not encontrados:
        return None, None
    from collections import Counter
    return Counter(encontrados).most_common(1)[0][0]


# -------------------------------------------------------------------------------
# ENCONTRAR ARQUIVOS
# -------------------------------------------------------------------------------

def encontrar_arquivos(pasta):
    """Classifica os arquivos da pasta materiais."""
    resultado = {
        "ml_evolucao":     None,
        "shopee_overview": None,
        "shopee_ads_csv":  None,
        "imagens":         [],
    }
    for arq in Path(pasta).iterdir():
        nome = arq.name.lower()
        ext  = arq.suffix.lower()
        if ext in (".jpg", ".jpeg", ".png"):
            resultado["imagens"].append(arq)
        elif "relatorio_evolucao" in nome or "relatorio_evolucao" in nome.replace("+", "_"):
            resultado["ml_evolucao"] = arq
        elif "sales_overview" in nome and ext in (".xlsx", ".xls"):
            resultado["shopee_overview"] = arq
        elif ("dados" in nome or "gerais" in nome) and ext == ".csv":
            resultado["shopee_ads_csv"] = arq
    return resultado


# -------------------------------------------------------------------------------
# PROCESSAMENTO AUTOMATICO
# -------------------------------------------------------------------------------

def processar_ml_evolucao(caminho):
    """Agrega dados diarios do relatorio de evolucao do ML."""
    print(f"\n  >> Lendo relatorio ML: {caminho.name}")
    wb = openpyxl.load_workbook(str(caminho), data_only=True)

    # Encontrar aba com "neg" no nome (Negocio)
    aba = None
    for nome in wb.sheetnames:
        if "neg" in nome.lower():
            aba = wb[nome]
            break
    if aba is None:
        aba = wb.active

    # Localizar linha de cabecalho (col A = "Data")
    header_row = None
    for r in range(1, 20):
        val = aba.cell(r, 1).value
        if val and str(val).strip().lower() == "data":
            header_row = r
            break
    if header_row is None:
        print("  [!] Cabecalho 'Data' nao encontrado na aba Negocio.")
        return None

    # Mapear colunas pelo nome
    col_map = {}
    for c in range(1, aba.max_column + 1):
        val = aba.cell(header_row, c).value
        if val:
            col_map[str(val).strip()] = c

    campos = {
        "Visitas":                "visitas",
        "Compradores únicos":     "compradores_unicos",
        "Novos compradores":      "novos_compradores",
        "Compradores existentes": "compradores_existentes",
        "Quantidade de vendas":   "qtd_vendas",
        "Unidades vendidas":      "unidades_vendidas",
        "Vendas brutas":          "vendas_brutas",
    }

    totais = {v: 0.0 for v in campos.values()}
    linhas = 0

    for r in range(header_row + 1, aba.max_row + 1):
        data_val = aba.cell(r, 1).value
        if not data_val:
            continue
        # Aceita qualquer linha nao vazia que nao seja o proprio cabecalho
        if str(data_val).strip().lower() in ("data", ""):
            continue
        for campo_orig, campo_dest in campos.items():
            col = col_map.get(campo_orig)
            if col:
                val = aba.cell(r, col).value
                if isinstance(val, (int, float)) and val == val:
                    totais[campo_dest] += val
                elif isinstance(val, str):
                    num = limpar_numero(val)
                    if num is not None:
                        totais[campo_dest] += num
        linhas += 1

    print(f"  [OK] {linhas} dias agregados")

    cu = totais["compradores_unicos"]
    vb = totais["vendas_brutas"]
    qv = totais["qtd_vendas"]
    uv = totais["unidades_vendidas"]
    vi = totais["visitas"]
    ce = totais["compradores_existentes"]

    totais["media_vendas_comprador"] = vb / cu if cu else 0
    totais["taxa_recompra"]          = ce / cu if cu else 0
    totais["conversao"]              = qv / vi if vi else 0
    totais["valor_medio_venda"]      = vb / qv if qv else 0
    totais["preco_medio_unidade"]    = vb / uv if uv else 0

    return totais


def processar_shopee_overview(caminho):
    """Le linha de resumo mensal do Shopee sales_overview."""
    print(f"\n  >> Lendo Shopee overview: {caminho.name}")
    wb = openpyxl.load_workbook(str(caminho), data_only=True)
    ws = wb.active

    # Linha 1 = cabecalho; Linha 2 = totais do periodo
    header_row = None
    data_row   = None
    for r in range(1, 10):
        val = ws.cell(r, 1).value
        if val and str(val).strip().lower() == "data":
            header_row = r
            for r2 in range(r + 1, r + 5):
                v2 = ws.cell(r2, 1).value
                if v2 is not None and str(v2).strip():
                    data_row = r2
                    break
            break

    if not header_row or not data_row:
        print("  [!] Estrutura do arquivo Shopee overview nao reconhecida.")
        return None

    col_map = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val:
            col_map[str(val).strip()] = c

    def get(nome):
        col = col_map.get(nome)
        if not col:
            return None
        val = ws.cell(data_row, col).value
        if isinstance(val, str):
            val = limpar_numero(val)
        return val

    def pct(v):
        """Garante que percentual esteja como decimal 0-1."""
        if v is None:
            return None
        if isinstance(v, str):
            v = limpar_numero(v)
            if v is None:
                return None
        return v / 100 if v > 1 else v

    resultado = {
        "visitantes":             get("Visitantes (Visitar)"),
        "num_pedidos":            get("Compradores (Pedidos Feitos)"),
        "pedidos_feitos_valor":   get("Vendas (Pedidos Feitos) (BRL)"),
        "pedidos_pagos_qtd":      get("Compradores (Pedidos Pagos)"),
        "pedidos_pagos_valor":    get("Vendas (Pedidos Pagos) (BRL)"),
        "ticket_medio":           get("Vendas por Comprador (Pedidos Pagos) (BRL)"),
        "taxa_conv_feitos":       pct(get("Taxa de Conversão (Visitados a Feitos)")),
        "taxa_conv_pagos":        pct(get("Taxa de Conversão (Visitados a Pagos)")),
        "taxa_conv_feitos_pagos": pct(get("Taxa de Conversão (Feitos a Pagos)")),
    }

    vp = resultado["pedidos_pagos_valor"] or 0
    vis = resultado["visitantes"] or 0
    print(f"  [OK] Visitantes: {vis:,.0f} | Vendas pagos: R${vp:,.2f}")
    return resultado


def processar_shopee_ads_csv(caminho):
    """Agrega dados de todos os grupos de anuncio do CSV Shopee."""
    print(f"\n  >> Lendo CSV anuncios Shopee: {caminho.name}")

    # Encontrar linha do cabecalho real (comeca com "#,")
    header_line = None
    with open(str(caminho), "r", encoding="utf-8") as f:
        for i, line in enumerate(f):
            if line.startswith("#,"):
                header_line = i
                break

    if header_line is None:
        print("  [!] Cabecalho do CSV nao encontrado.")
        return None

    df = pd.read_csv(str(caminho), skiprows=header_line, encoding="utf-8")

    # Manter apenas linhas de dados (# = numero inteiro)
    df = df[df["#"].apply(lambda x: str(x).strip().isdigit())].copy()

    cols_num = [
        "Impressoes", "Cliques", "Conversoes", "Conversoes Diretas",
        "Itens Vendidos", "Itens Vendidos Diretos", "GMV",
        "Receita direta", "Despesas",
    ]
    # Mapeamento de nomes com acento -> sem acento (como estao no CSV)
    cols_csv = {
        "Impressoes":          "Impressões",
        "Cliques":             "Cliques",
        "Conversoes":          "Conversões",
        "Conversoes Diretas":  "Conversões Diretas",
        "Itens Vendidos":      "Itens Vendidos",
        "Itens Vendidos Diretos": "Itens Vendidos Diretos",
        "GMV":                 "GMV",
        "Receita direta":      "Receita direta",
        "Despesas":            "Despesas",
    }

    def soma(col_sem_acento):
        col_com_acento = cols_csv[col_sem_acento]
        if col_com_acento in df.columns:
            return pd.to_numeric(df[col_com_acento], errors="coerce").fillna(0).sum()
        return 0.0

    imp  = soma("Impressoes")
    clq  = soma("Cliques")
    conv = soma("Conversoes")
    cvd  = soma("Conversoes Diretas")
    itv  = soma("Itens Vendidos")
    ivd  = soma("Itens Vendidos Diretos")
    gmv  = soma("GMV")
    rec  = soma("Receita direta")
    desp = soma("Despesas")

    totais = {
        "pub_impressoes":       imp,
        "pub_cliques":          clq,
        "pub_ctr":              clq / imp  if imp  else 0,
        "pub_conversoes":       conv,
        "pub_conv_diretas":     cvd,
        "pub_taxa_conv":        conv / clq if clq  else 0,
        "pub_taxa_conv_direta": cvd  / clq if clq  else 0,
        "pub_itens_vendidos":   itv,
        "pub_itens_diretos":    ivd,
        "pub_gmv":              gmv,
        "pub_receita_ads":      rec,
        "pub_despesas":         desp,
        "pub_roas":             gmv  / desp if desp else 0,
        "pub_roas_direto":      rec  / desp if desp else 0,
        "pub_acos":             desp / gmv  if gmv  else 0,
        "pub_acos_direto":      desp / rec  if rec  else 0,
    }

    print(f"  [OK] {len(df)} grupos | Despesas: R${desp:,.2f} | GMV: R${gmv:,.2f}")
    return totais


# -------------------------------------------------------------------------------
# COLETA INTERATIVA DE IMAGENS
# -------------------------------------------------------------------------------

def _coletar(titulo, campos):
    print(f"\n{'='*62}")
    print(f"  {titulo}")
    print(f"{'='*62}")
    dados = {}
    for chave, label in campos:
        dados[chave] = input_numero(f"  {label}: ")
    return dados


def coletar_ml_ads(img):
    print(f"\n  Abrindo imagem: {img.name}")
    abrir_imagem(img)
    return _coletar("MERCADO LIVRE -- PUBLICIDADE (imagem com grafico azul)", [
        ("pub_impressoes",          "Impressoes"),
        ("pub_cliques",             "Cliques"),
        ("pub_vendas_product_ads",  "Vendas por Product Ads"),
        ("pub_vendas_sem_products", "Vendas sem Product Ads"),
        ("pub_investimento",        "Investimento (R$)"),
        ("pub_receita",             "Receita (R$)"),
    ])


def coletar_ml_afiliados(img):
    print(f"\n  Abrindo imagem: {img.name}")
    abrir_imagem(img)
    return _coletar("MERCADO LIVRE -- AFILIADOS (aba Rendimento geral)", [
        ("af_receita",    "Vendas / Receita Afiliados (R$)"),
        ("af_unidades",   "Unidades vendidas"),
        ("af_qtd_vendas", "Quantidade de vendas"),
        ("af_custo",      "Custo estimado (R$)"),
    ])


def coletar_shopee_afiliados(img):
    print(f"\n  Abrindo imagem: {img.name}")
    abrir_imagem(img)
    print("  [!] Valores como 'R$70mil' -> digite: 70000")
    return _coletar("SHOPEE -- AFILIADOS (Principais Indicadores)", [
        ("af_vendas",             "Vendas (R$)"),
        ("af_itens_brutos",       "Itens vendidos brutos"),
        ("af_pedidos",            "Pedidos"),
        ("af_cliques",            "Cliques"),
        ("af_comissao",           "Comissao Estimada (R$)"),
        ("af_roi",                "ROI"),
        ("af_compradores_totais", "Compradores totais"),
        ("af_novos_compradores",  "Novos compradores"),
    ])


def coletar_site():
    print(f"\n{'='*62}")
    print("  SITE -- Dados Gerais")
    print(f"{'='*62}")
    resp = input("  Deseja inserir dados do Site? [s/N]: ").strip().lower()
    if resp != "s":
        return None
    return _coletar("SITE -- Dados do mes", [
        ("usuarios",        "Usuarios"),
        ("num_pedidos",     "N. Pedidos"),
        ("valor_total",     "Valor Total (R$)"),
        ("ticket_medio",    "Ticket Medio (R$)"),
        ("valor_investido", "Valor Investido Google + Meta (R$)"),
        ("google",          "Google Ads (R$)"),
        ("meta_ads",        "Meta Ads (R$)"),
        ("taxa_tacos",      "Taxa TACOS (ex: 0.21 para 21%)"),
        ("taxa_conversao",  "Taxa de Conversao (ex: 0.0092 para 0.92%)"),
    ])


# -------------------------------------------------------------------------------
# ATUALIZACAO DA PLANILHA
# -------------------------------------------------------------------------------

def encontrar_linha_mes(ws, serial_mes, linha_inicio):
    """Retorna a linha do mes ou None se nao encontrada."""
    from datetime import timedelta
    alvo = date(1899, 12, 30) + timedelta(days=int(serial_mes))

    for r in range(linha_inicio, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val is None:
            continue
        # Serial numerico (como a planilha armazena)
        if isinstance(val, (int, float)):
            d = date(1899, 12, 30) + timedelta(days=int(val))
            if d.year == alvo.year and d.month == alvo.month:
                return r
        # Datetime (caso openpyxl converta automaticamente)
        elif isinstance(val, (date, datetime)):
            d = val if isinstance(val, date) else val.date()
            if d.year == alvo.year and d.month == alvo.month:
                return r
    return None


def ultima_linha_dados(ws, linha_inicio):
    """Ultima linha que tem valor na coluna 1."""
    ultima = linha_inicio - 1
    for r in range(linha_inicio, ws.max_row + 1):
        if ws.cell(r, 1).value not in (None, ""):
            ultima = r
    return ultima


def atualizar_aba(ws, col_map, formula_cols, dados, ano, mes, linha_inicio):
    """Escreve (ou atualiza) a linha do mes na aba."""
    serial = mes_serial_excel(ano, mes)

    linha = encontrar_linha_mes(ws, serial, linha_inicio)
    ult   = ultima_linha_dados(ws, linha_inicio)

    if linha is None:
        linha = ult + 1
        print(f"  -> Nova linha: {linha}")
    else:
        print(f"  -> Linha existente: {linha} (atualizacao)")

    # Copiar estilo e formulas da linha acima
    linha_ref = linha - 1
    if linha_ref >= linha_inicio:
        for c in range(1, ws.max_column + 1):
            copiar_estilo(ws.cell(linha_ref, c), ws.cell(linha, c))
        for c in formula_cols:
            formula_acima = ws.cell(linha_ref, c).value
            if formula_acima and str(formula_acima).startswith("="):
                ws.cell(linha, c).value = ajustar_formula(formula_acima, delta=1)

    # Gravar o serial do mes
    ws.cell(linha, 1).value = serial

    # Gravar valores
    for campo, col in col_map.items():
        if campo == "mes" or col in formula_cols:
            continue
        val = dados.get(campo)
        if val is not None:
            ws.cell(linha, col).value = val

    return linha


# -------------------------------------------------------------------------------
# RESUMO
# -------------------------------------------------------------------------------

def mostrar_resumo(titulo, dados, col_map):
    print(f"\n  +-- {titulo}")
    for campo, val in dados.items():
        if val is None:
            continue
        col = col_map.get(campo, "?")
        if isinstance(val, float):
            if val > 1000:
                s = f"{val:,.2f}"
            elif val <= 1:
                s = f"{val:.4f}  ({val*100:.2f}%)"
            else:
                s = f"{val:.4f}"
        else:
            s = str(val)
        print(f"  |  C{col:02d}  {campo}: {s}")
    print("  +--")


# -------------------------------------------------------------------------------
# MAIN
# -------------------------------------------------------------------------------

def main():
    print()
    print("=" * 62)
    print("   ATUALIZADOR -- PERFORMANCE MBS MKTPLACE")
    print("=" * 62)

    # 1. Detectar mes
    print(f"\nPasta materiais: {MATERIAIS_DIR}")
    ano, mes = detectar_mes(MATERIAIS_DIR)

    if ano and mes:
        nome_mes = f"{MESES_PT[mes]}/{str(ano)[2:]}"
        print(f"\n[OK] Mes detectado: {nome_mes}  ({mes:02d}/{ano})")
        resp = input("     Correto? [S/n]: ").strip().lower()
        if resp == "n":
            ano = int(input("     Ano (ex: 2026): "))
            mes = int(input("     Mes 1-12: "))
            nome_mes = f"{MESES_PT[mes]}/{str(ano)[2:]}"
    else:
        print("\n[!] Mes nao detectado nos nomes dos arquivos.")
        ano = int(input("    Ano (ex: 2026): "))
        mes = int(input("    Mes 1-12: "))
        nome_mes = f"{MESES_PT[mes]}/{str(ano)[2:]}"

    print(f"\nMes de referencia: {nome_mes}")

    # 2. Localizar arquivos
    arqs = encontrar_arquivos(MATERIAIS_DIR)
    dados_ml     = {}
    dados_shopee = {}

    # 3. ML Evolution (automatico)
    if arqs["ml_evolucao"]:
        r = processar_ml_evolucao(arqs["ml_evolucao"])
        if r:
            dados_ml.update(r)
    else:
        print("\n[!] Relatorio ML evolucao nao encontrado em materiais.")

    # 4. Shopee overview (automatico)
    if arqs["shopee_overview"]:
        r = processar_shopee_overview(arqs["shopee_overview"])
        if r:
            dados_shopee.update(r)
    else:
        print("\n[!] Shopee sales_overview nao encontrado em materiais.")

    # 5. Shopee ads CSV (automatico)
    if arqs["shopee_ads_csv"]:
        r = processar_shopee_ads_csv(arqs["shopee_ads_csv"])
        if r:
            dados_shopee.update(r)
    else:
        print("\n[!] CSV de anuncios Shopee nao encontrado em materiais.")

    # 6. Imagens (IA automatica)
    imagens = arqs["imagens"]
    if imagens:
        api_key = carregar_api_key()
        print(f"\n{'='*62}")
        print(f"  IMAGENS ({len(imagens)} encontradas) -- leitura automatica com IA")
        print(f"{'='*62}")

        img_ml, img_sh = processar_imagens_ia(imagens, api_key)
        dados_ml.update(img_ml)
        dados_shopee.update(img_sh)

    # 7. Site (interativo)
    dados_site = coletar_site()

    # 8. Resumo e confirmacao
    print(f"\n{'='*62}")
    print(f"  RESUMO -- {nome_mes.upper()}")
    print(f"{'='*62}")
    if dados_ml:
        mostrar_resumo("Mercado Livre", dados_ml, ML_COLS)
    if dados_shopee:
        mostrar_resumo("Shopee", dados_shopee, SHOPEE_COLS)
    if dados_site:
        mostrar_resumo("Site", dados_site, SITE_COLS)

    print(f"\nArquivo alvo: {PERFORMANCE_FILE}")
    resp = input("\nConfirmar gravacao? [S/n]: ").strip().lower()
    if resp == "n":
        print("\nOperacao cancelada.")
        return

    # 9. Backup
    bkp = Path(PERFORMANCE_FILE).with_suffix(f".bkp_{ano}{mes:02d}.xlsx")
    shutil.copy2(PERFORMANCE_FILE, bkp)
    print(f"\n[OK] Backup: {bkp.name}")

    # 10. Gravar
    print("     Abrindo planilha...")
    wb = openpyxl.load_workbook(PERFORMANCE_FILE)

    if dados_ml:
        ws = wb["Mercado Livre"]
        ln = atualizar_aba(ws, ML_COLS, ML_FORMULA_COLS, dados_ml, ano, mes, ML_LINHA_DADOS)
        print(f"[OK] Aba Mercado Livre -> linha {ln}")

    if dados_shopee:
        ws = wb["Shopee"]
        ln = atualizar_aba(ws, SHOPEE_COLS, SHOPEE_FORMULA_COLS, dados_shopee, ano, mes, SHOPEE_LINHA_DADOS)
        print(f"[OK] Aba Shopee -> linha {ln}")

    if dados_site:
        ws = wb["Site"]
        ln = atualizar_aba(ws, SITE_COLS, SITE_FORMULA_COLS, dados_site, ano, mes, SITE_LINHA_DADOS)
        print(f"[OK] Aba Site -> linha {ln}")

    wb.save(PERFORMANCE_FILE)
    print(f"\n[OK] Planilha salva com sucesso!")
    print(f"     {PERFORMANCE_FILE}\n")


if __name__ == "__main__":
    main()
